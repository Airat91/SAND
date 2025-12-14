import project_generator as generator
import colorama
from colorama import Fore, Back, Style, init
import openpyxl
import sand_reg
import copy

def sand_reg_h_processing(Proj):
    """
    Insert sand_prop_xxx_t structs from sand_reg.py to sand_reg.h
    :param Proj: class Project (project_generator.Project)
    :return:
    """

    step = "1 Read sand_reg.py and find structs"
    for prop_name in sand_reg.sand_prop_list:
        property = sand_reg.sand_prop_list[prop_name]
        if "header" in property:
            if property["header"]["type"] != "sand_header_t":
                print(Fore.YELLOW + "WARNING: generator/sand_reg.py: '{}' header is '{}' (expected 'sand_header_t')"
                                    "\n".format(prop_name, property["header"]["type"]) + Style.RESET_ALL)
            else:
                Proj.sand_properties["prop_name"].append(prop_name)
                Proj.sand_properties[prop_name] = property
        else:
            Proj.errors["err_msg"].append("generator/sand_reg.py: '{}' haven't header".format(prop_name))
            Proj.errors["err_cnt"] += 1

    file_name = "sand_reg_h"
    file = open(Proj.path[file_name], "r", encoding='UTF-8')
    text_lines = file.readlines()
    file.close()
    # find start and end of regs_description struct
    start_insert = generator.get_msg_line_nmbr(Proj,file_name, "sand_properties", "insert_start")
    end_insert = generator.get_msg_line_nmbr(Proj,file_name, "sand_properties", "insert_end")
    generator_insertion_find = False
    if start_insert != "not_found" and end_insert != "not_found":
        if start_insert < end_insert:
            # We found end of regs_description struct
            generator_insertion_find = True
    if generator_insertion_find == False:
        Proj.errors["err_msg"].append("Don't found #generator_message in " + file_name)
        Proj.errors["err_cnt"] += 1
    else:
        step = "1. Add sand_properties to buffer for not corrupting file if errors"
        buffer = []
        try:
            step = "1.1 Write generator message"
            buffer.append("// This part of file generated automatically, don't change it\n\n")

            step = "1.2 Calc words len for align text"
            max_spaces = [0,0]
            for param in sand_reg.sand_header_t:
                if len(sand_reg.sand_header_t[param]["type"]) > max_spaces[0]:
                    max_spaces[0] = len(sand_reg.sand_header_t[param]["type"])
                if len(param) > max_spaces[1]:
                    max_spaces[1] = len(param)
            for prop_name in Proj.sand_properties["prop_name"]:
                property = Proj.sand_properties[prop_name]
                for param in property:
                    if len(property[param]["type"]) > max_spaces[0]:
                        max_spaces[0] = len(property[param]["type"])
                    if len(param) > max_spaces[1]:
                        max_spaces[1] = len(param)
            for var_type in sand_reg.sand_var_t:
                if len(var_type) > max_spaces[0]:
                    max_spaces[0] = len(var_type)
                if len(sand_reg.sand_var_t[var_type]["comment"]) > max_spaces[1]:
                    max_spaces[1] = len(sand_reg.sand_var_t[var_type]["comment"])
            for acc_lvl in sand_reg.sand_access_lvl_t:
                if len(acc_lvl) > max_spaces[0]:
                    max_spaces[0] = len(acc_lvl)
                if len(str(sand_reg.sand_access_lvl_t[acc_lvl]["value"])) > max_spaces[1]:
                    max_spaces[1] = len(str(sand_reg.sand_access_lvl_t[acc_lvl]["value"]))

            step = "1.3 Write sand_prop enumeration"
            buffer.append("typedef enum{\n")
            for prop_name in Proj.sand_properties["prop_name"]:
                buffer.append("\t{},\n".format(prop_name.replace("_t","").upper()))
            buffer.append("}sand_prop_enum_t;\n\n")

            step = "1.4 Write sand_var enumeration"
            buffer.append("typedef enum{\n")
            for var_type in sand_reg.sand_var_t:
                space_0 = " "*(max_spaces[0] - len(var_type) + max_spaces[1] + 2)
                space_1 = " "*(max_spaces[1] - len(sand_reg.sand_var_t[var_type]["comment"]) + 1)
                buffer.append("\t{},{}// {}{}({} byte)\n".format(var_type.upper(), space_0, sand_reg.sand_var_t[var_type]["comment"],
                                                                 space_1, sand_reg.sand_var_t[var_type]["byte_num"]))
            buffer.append("}sand_var_t;\n\n")

            step = "1.4 Write sand_access_lvl enumeration"
            buffer.append("typedef enum{\n")
            for acc_lvl in sand_reg.sand_access_lvl_t:
                space_0 = " "*(max_spaces[0] - len(acc_lvl) + 1)
                space_1 = " "*(max_spaces[1] - len(str(sand_reg.sand_access_lvl_t[acc_lvl]["value"])) - 1)
                buffer.append("\t{}{}= {},{}// {}\n".format(acc_lvl.upper(), space_0, sand_reg.sand_access_lvl_t[acc_lvl]["value"],
                                                                 space_1, sand_reg.sand_access_lvl_t[acc_lvl]["comment"]))
            buffer.append("}sand_access_lvl;\n\n")

            step = "1.5 Write header to buffer"
            buffer.append("typedef struct{\n")
            for param in sand_reg.sand_header_t:
                space_0 = " "*(max_spaces[0] - len(sand_reg.sand_header_t[param]["type"]) + 1)
                space_1 = " "*(max_spaces[1] - len(param) + 1)
                buffer.append("\t{}{}{};{}// {}\n".format(sand_reg.sand_header_t[param]["type"], space_0, param,
                                                          space_1, sand_reg.sand_header_t[param]["comment"]))
            buffer.append("}sand_header_t;\n\n")

            step = "1.6 Write sand_prop_t to buffer"
            for prop_name in Proj.sand_properties["prop_name"]:
                property = Proj.sand_properties[prop_name]
                buffer.append("typedef struct{\n")
                for param in property:
                    space_0 = " "*(max_spaces[0] - len(property[param]["type"]) + 1)
                    space_1 = " "*(max_spaces[1] - len(param) + 1)
                    buffer.append("\t{}{}{};{}// {}\n".format(property[param]["type"], space_0, param,
                                                              space_1, property[param]["comment"]))
                buffer.append("}"+"{};\n\n".format(prop_name))
        except:
            Proj.errors["err_msg"].append("Error during adding property lists to " + file_name + ", step: " + step)
            Proj.errors["err_cnt"] += 1

    step = "4. Rewrite file with new insertion from generator"
    if generator_insertion_find == True:
        step = "2. Rewrite sand_reg.h file with new insertion from generator"
        file = open(Proj.path[file_name], 'w', encoding='UTF-8')
        line_index = 0
        step = "3. Write file before insertion"
        while line_index <= start_insert:
            file.writelines(text_lines[line_index])
            line_index += 1
        line_index -= 1
        step = "4. Write insertion"
        if len(buffer) > 0:
            for line in buffer:
                file.writelines(line)
        step = "5. Write file afer insertion"
        line_index = end_insert - 1
        while line_index < len(text_lines):
            file.writelines(text_lines[line_index])
            line_index += 1
        step = "6. Close modified file"
        file.close()

def reg_map_module_xls_processing(Proj):
    """
    Read regs from reg_map_module.xls to Proj
    1. Read structs into Proj[struct_list]
    2. Read regs into Proj[struct_list][struct][reg_list]
    3. Read properties into Proj[prop_list]
    4. Add regs into Proj[prop_list][prop][reg_list]
    :param Proj: class Project (project_generator.Project)
    :return:
    """
    if len(Proj.struct_list) > 0:
        return
    file_name = Proj.path["reg_map_{}_xls".format(Proj.module)]
    file_xls = openpyxl.load_workbook(file_name)
    # Check struct_list existence in file
    if "struct_list" not in file_xls.sheetnames:
        Proj.errors["err_msg"].append("\"reg_map_{}.xls\": struct list not found".format(Proj.module))
        Proj.errors["err_cnt"] += 1
    # Read struct_list and check structs existence in file
    for worksheet in file_xls.worksheets:
        if worksheet.title == "struct_list":
            file_xls.active = worksheet
            break
    found = False
    for column in range(worksheet.min_column, worksheet.max_column + 1):
        if worksheet.cell(1, column).value == "Struct name":
            found = True
            break
    if found == False:
        Proj.errors["err_msg"].append("\"reg_map_{}.xls\": \"Struct name\" column not found".format(Proj.module))
        Proj.errors["err_cnt"] += 1
    struct_list_xls = []
    for row in range (2, worksheet.max_row + 1):
        struct_list_xls.append(worksheet.cell(row, column).value)
    for struct in struct_list_xls:
        if struct not in file_xls.sheetnames:
            print(Fore.YELLOW + "WARNING: \"reg_map_{}.xls\": struct \"{}\" define in struct_list but don't found in "
                                "sheets\nStruct \"{}\" excluded from project".format(Proj.module, struct, struct)
                  + Style.RESET_ALL)
            struct_list_xls.remove(struct)
    # Add properties to project
    for prop in sand_reg.sand_prop_list:
        Proj.prop_list[prop] = {
            "reg_list" : {}
        }
    # Add structs to project
    for struct_name in struct_list_xls:
        Proj.struct_list[struct_name] = {
            "name": struct_name,
            "reg_list": {},
            "byte_size": 0,
        }
        project_struct = Proj.struct_list[struct_name]
        # Find sheet of current struct
        for worksheet in file_xls.worksheets:
            if worksheet.title == struct_name:
                break
        # Read properties in first row
        property_headers = {}   # column indexes for property parameters
        property_list = {}      # list of properties founded on sheet
        column = 1
        max_column = worksheet.max_column
        while column < (max_column + 1):
            property = worksheet.cell(1, column).value
            if property != None:
                if property == "SAND_PROP_END":
                    break;
                property = property.lower() + "_t"
                if property in sand_reg.sand_prop_list:
                    property_list[property] = copy.copy(sand_reg.sand_prop_list[property])
                    property_list[property]["is_exist"] = False
                    # Read property parameters and save its indexes into property_headers
                    param = worksheet.cell(2, column).value
                    if param not in property_list[property]:
                        avl_list = list(property_list[property].keys())
                        avl_list.remove("header")
                        print(Fore.YELLOW + "WARNING: \"reg_map_{}.xls\" struct \"{}\" property \"{}\" param \"{}\" is "
                                            "unknown.\nAvailable parameters defined in \"sand_reg.py\": {}".format(
                            Proj.module, struct_name, property, param, avl_list) + Style.RESET_ALL)
                    else:
                        property_headers[column] = {
                            "property": property,
                            "param": param,
                        }
                    column += 1
                    while ((worksheet.cell(1, column).value == None) and (column < (max_column + 1))):
                        param = worksheet.cell(2, column).value
                        if param not in property_list[property]:
                            avl_list = list(property_list[property].keys())
                            avl_list.remove("header")
                            print(
                                Fore.YELLOW + "WARNING: \"reg_map_{}.xls\" struct \"{}\" property \"{}\" param \"{}\" "
                                              "is unknown.\nAvailable parameters defined in \"sand_reg.py\": {}".format(
                                    Proj.module, struct_name, property, param, avl_list) + Style.RESET_ALL)
                        else:
                            property_headers[column] = {
                                "property": property,
                                "param": param,
                            }
                        column += 1
                    column -= 1
            column += 1
        # Read regs on sheet
        for row in range(3, worksheet.max_row + 1):
            reg = sand_reg.reg()
            for column in range(1, max_column):
                property = property_headers[column]["property"]
                param = property_headers[column]["param"]
                reg.prop_list[property][param]["value"] = worksheet.cell(row, column).value
                if worksheet.cell(row, column).value != None:
                    reg.prop_list[property]["is_exist"] = True
            reg_name = reg.prop_list["sand_prop_base_t"]["name"]["value"]
            reg.name = reg_name
            reg.type = reg.prop_list["sand_prop_base_t"]["type"]["value"]
            reg.struct = struct_name
            Proj.struct_list[struct_name]["reg_list"][reg_name] = reg
        print(Fore.GREEN + "Struct \"{}\" found {} registers".format(struct_name,
                                                                     len(Proj.struct_list[struct_name]["reg_list"]))
              + Style.RESET_ALL)
        for reg_name in Proj.struct_list[struct_name]["reg_list"]:
            reg = Proj.struct_list[struct_name]["reg_list"][reg_name]
            for property_name in reg.prop_list:
                property = reg.prop_list[property_name]
                # Add regs to Proj.prop_lists
                if "is_exist" in property:
                    if property["is_exist"] == True:
                        #reg["struct_name"] = struct_name
                        Proj.prop_list[property_name]["reg_list"][reg_name] = reg
    print(Fore.GREEN + "reg_map_module_xls_processing done" + Style.RESET_ALL)

def regs_module_h_processing(Proj):
    """
    Write project structs into regs_module.h
    :param Proj: class Project (project_generator.Project)
    :return:
    """
    file_name = "regs_module_h"
    file = open(Proj.path[file_name], "r", encoding='UTF-8')
    text_lines = file.readlines()
    file.close()

    step = "1 Find start and end of sand_struct_define"
    define_insert_start = generator.get_msg_line_nmbr(Proj, file_name, "sand_struct_define", "insert_start")
    define_insert_end = generator.get_msg_line_nmbr(Proj, file_name, "sand_struct_define", "insert_end")
    define_insert_find = False
    if define_insert_start != "not_found" and define_insert_end != "not_found":
        if define_insert_start < define_insert_end:
            # We found end of regs_description struct
            define_insert_find = True
    if define_insert_find == False:
        Proj.errors["err_msg"].append("Don't found #generator_message \"sand_struct_define\" in " + file_name)
        Proj.errors["err_cnt"] += 1
    else:

        step = "1.1. Add defines to buffer for not corrupting file if errors"
        buffer_define = []
        try:

            step = "1.1.1 Write generator message"
            buffer_define.append("// This part of file generated automatically, don't change it\n")

            step = "1.1.2 Calc words len for align text"
            max_spaces = [0]
            for struct_name in Proj.struct_list:
                define_str = "{}_STRUCT_SIZE".format(struct_name)
                if len(define_str) > max_spaces[0]:
                    max_spaces[0] = len(define_str)

            for prop_name in Proj.prop_list:
                define_str = "{}_REG_NUM".format(prop_name.replace("_t",""))
                if len(define_str) > max_spaces[0]:
                    max_spaces[0] = len(define_str)

            step = "1.1.3 Write device name define"
            define_str = "DEVICE_NAME"
            if len(define_str) > max_spaces[0]:
                max_spaces[0] = len(define_str)
            space_0 = " " * (max_spaces[0] - len(define_str) + 1)
            buffer_define.append("#define {}{}\t\"{}\"\n".format(define_str, space_0, Proj.name))
            buffer_define.append("\n")

            step = "1.1.4 Write save_data_size"
            define_str = "SAND_SAVE_DATA_SIZE"
            if len(define_str) > max_spaces[0]:
                max_spaces[0] = len(define_str)
            space_0 = " " * (max_spaces[0] - len(define_str) + 1)
            save_info = Proj.prop_list["sand_prop_save_t"]["save_info"]
            buffer_define.append("#define {}{}\t{}\n".format(define_str, space_0, save_info["save_area_size"]))
            buffer_define.append("\n")

            step = "1.1.5 Write defines to buffer_define"
            for struct_name in Proj.struct_list:
                struct = Proj.struct_list[struct_name]
                struct_byte_size = struct["byte_size"]
                define_str = "{}_STRUCT_SIZE".format(struct_name).upper()
                space_0 = " " * (max_spaces[0] - len(define_str) + 1)
                buffer_define.append("#define {}{}\t{}\n".format(define_str, space_0, struct_byte_size))

            buffer_define.append("\n")
            for prop_name in Proj.prop_list:
                prop = Proj.prop_list[prop_name]
                prop_reg_num = len(prop["reg_list"])
                define_str = "{}_REG_NUM".format(prop_name.replace("_t","")).upper()
                space_0 = " " * (max_spaces[0] - len(define_str) + 1)
                buffer_define.append("#define {}{}\t{}\n".format(define_str, space_0, prop_reg_num))

        except:
            Proj.errors["err_msg"].append("Error during adding property lists to " + file_name + ", step: " + step)
            Proj.errors["err_cnt"] += 1

    # 2. Find start and end of sand_struct
    struct_insert_start = generator.get_msg_line_nmbr(Proj, file_name, "sand_struct", "insert_start")
    struct_insert_end = generator.get_msg_line_nmbr(Proj, file_name, "sand_struct", "insert_end")
    struct_insertion_find = False
    if struct_insert_start != "not_found" and struct_insert_end != "not_found":
        if struct_insert_start < struct_insert_end:
            # We found end of sand_struct
            struct_insertion_find = True
    if struct_insertion_find == False:
        Proj.errors["err_msg"].append("Don't found #generator_message \"sand_struct\" in " + file_name)
        Proj.errors["err_cnt"] += 1
    else:
        # 2.1. Add structs to buffer for not corrupting file if errors
        buffer_struct = []
        try:
            # 2.1.1 Write generator message
            buffer_struct.append("// This part of file generated automatically, don't change it\n")

            # 2.1.2 Calc words len for align text
            max_spaces = [0, 0, 0, 0]
            for struct_name in Proj.struct_list:
                struct = Proj.struct_list[struct_name]
                for reg_name in struct["reg_list"]:
                    reg = struct["reg_list"][reg_name]
                    # reg_type
                    if len(reg.type) > max_spaces[0]:
                        max_spaces[0] = len(reg.type)
                    #reg_name
                    if len(reg.name) > max_spaces[1]:
                        max_spaces[1] = len(reg.name)
                    #array_len
                    array_len = reg.prop_list["sand_prop_base_t"]["array_len"]["value"]
                    if array_len > 1:
                        if len("[{}]".format(array_len)) > max_spaces[2]:
                            max_spaces[2] = len("[{}]".format(array_len))
            # 2.1.3 Write struct to buffer_struct
            for struct_name in Proj.struct_list:
                struct = Proj.struct_list[struct_name]
                buffer_struct.append("typedef union{\n\tstruct MCU_PACK{\n")
                for reg_name in struct["reg_list"]:
                    reg = struct["reg_list"][reg_name]
                    space_0 = " " * (max_spaces[0] - len(reg.type) + 1)
                    space_1 = " " * (max_spaces[1] - len(reg.name) + 1)
                    array_len = reg.prop_list["sand_prop_base_t"]["array_len"]["value"]
                    array_len_str = ""
                    if array_len > 1:
                        array_len_str = "[{}]".format(array_len)
                        space_2 = " " * (max_spaces[2] - len(array_len_str))
                    else:
                        space_2 = " " * max_spaces[2]
                    descr = reg.prop_list["sand_prop_base_t"]["description"]["value"]
                    buffer_struct.append("\t\t{}{}{}{};{}{}// {}\n".format(reg.type, space_0, reg.name, array_len_str, space_1,
                                                                space_2, descr))
                buffer_struct.append("\t}vars;\n")
                struct_size = struct_name.upper()+"_STRUCT_SIZE"
                buffer_struct.append("\tu8 bytes[{}];\n".format(struct_size))
                buffer_struct.append("}"+"{}_struct;\n\n".format(struct_name))
        except:
            Proj.errors["err_msg"].append("Error during adding structs to " + file_name)
            Proj.errors["err_cnt"] += 1

    # 3. Find start and end of sand_struct_external
    external_insert_start = generator.get_msg_line_nmbr(Proj, file_name, "sand_struct_external", "insert_start")
    external_insert_end = generator.get_msg_line_nmbr(Proj, file_name, "sand_struct_external", "insert_end")
    external_insertion_find = False
    if external_insert_start != "not_found" and external_insert_end != "not_found":
        if external_insert_start < external_insert_end:
            # We found end of sand_struct_external
            external_insertion_find = True
    if external_insertion_find == False:
        Proj.errors["err_msg"].append("Don't found #generator_message \"sand_struct_external\" in " + file_name)
        Proj.errors["err_cnt"] += 1
    else:
        # 3.1. Add structs to buffer for not corrupting file if errors
        buffer_external = []
        try:
            # 3.1.1 Write generator message
            buffer_external.append("// This part of file generated automatically, don't change it\n")

            # 3.1.2 Calc words len for align text
            max_spaces = [0]
            for struct_name in Proj.struct_list:
                ext_str = "{}_struct".format(struct_name)
                if len(ext_str) > max_spaces[0]:
                    max_spaces[0] = len(ext_str)
            for prop_name in Proj.prop_list:
                ext_str = "const {}".format(prop_name)
                if len(ext_str) > max_spaces[0]:
                    max_spaces[0] = len(ext_str)

            # 3.1.3 Write external variables declaration to buffer_external
            for struct_name in Proj.struct_list:
                ext_str = "{}_struct".format(struct_name)
                space_0 = " " * (max_spaces[0] - len(ext_str) + 1)
                buffer_external.append("extern {}{}\t{};\n".format(ext_str, space_0, struct_name))
            buffer_external.append("\n")

            for prop_name in Proj.prop_list:
                ext_str = "const {}".format(prop_name)
                prop_list_name = prop_name.replace("_t", "_list")
                space_0 = " " * (max_spaces[0] - len(ext_str) + 1)
                buffer_external.append("extern {}{}\t{}[];\n".format(ext_str, space_0, prop_list_name))
        except:
            Proj.errors["err_msg"].append("Error during adding external variables declaration to " + file_name)
            Proj.errors["err_cnt"] += 1

    # 4. Rewrite file with new insertion from generator
    if define_insert_find == True and struct_insertion_find == True and external_insertion_find == True:
        # 4.1 Check sequence of insertions
        if (define_insert_start > struct_insert_start) or (define_insert_end > external_insert_start) or (struct_insert_start > external_insert_start):
            Proj.errors["err_msg"].append("Error sequence of insertions in " + file_name)
            Proj.errors["err_cnt"] += 1
            return False

        file = open(Proj.path[file_name], 'w', encoding='UTF-8')
        line_index = 0
        # 4.1. Write file before insertion
        while line_index <= define_insert_start:
            file.writelines(text_lines[line_index])
            line_index += 1
        # 4.2. Write defines insertion
        for line in buffer_define:
            file.writelines(line)
        # 4.3. Write file between defines and structs insertion
        line_index = define_insert_end
        while line_index <= struct_insert_start:
            file.writelines(text_lines[line_index])
            line_index += 1
        #line_index -= 1
        # 4.4. Write structs insertion
        for line in buffer_struct:
            file.writelines(line)
        # 4.5. Write file between structs and external insertion
        line_index = struct_insert_end
        while line_index <= external_insert_start:
            file.writelines(text_lines[line_index])
            line_index += 1
        #line_index -= 1
        # 4.6. Write external insertion
        for line in buffer_external:
            file.writelines(line)
        # 4.7. Write file after external insertion
        line_index = external_insert_end
        while line_index < len(text_lines):
            file.writelines(text_lines[line_index])
            line_index += 1
        # 4.8. Close modified file
        file.close()

    return True

def regs_module_c_processing(Proj):
    """
    Write project regs declarations into regs_module.c
    :param Proj: class Project (project_generator.Project)
    :return:
    """
    step = "Open file"
    file_name = "regs_module_c"
    file = open(Proj.path[file_name], "r", encoding='UTF-8')
    text_lines = file.readlines()
    file.close()

    step = "1 Find start and end of sand_properties"
    prop_insert_start = generator.get_msg_line_nmbr(Proj, file_name, "sand_properties", "insert_start")
    prop_insert_end = generator.get_msg_line_nmbr(Proj, file_name, "sand_properties", "insert_end")
    prop_insert_find = False
    if prop_insert_start != "not_found" and prop_insert_end != "not_found":
        if prop_insert_start < prop_insert_end:
            # We found end of regs_description struct
            prop_insert_find = True
    if prop_insert_find == False:
        Proj.errors["err_msg"].append("Don't found #generator_message \"sand_properties\" in " + file_name)
        Proj.errors["err_cnt"] += 1
    else:
        step = "1.1. Add defines to buffer for not corrupting file if errors"
        buffer_prop = []
        try:

            step = "1.1.1 Write generator message"
            buffer_prop.append("// This part of file generated automatically, don't change it\n\n")

            step = "1.1.2 Write global structs declaration"
            max_spaces = [0] * 2

            step = "1.1.2.1 Calc words len for align text"
            for struct_name in Proj.struct_list:
                struct_declaration = struct_name + "_struct"
                if len(struct_declaration) > max_spaces[0]:
                    max_spaces[0] = len(struct_declaration)
                if len(struct_name) > max_spaces[1]:
                    max_spaces[1] = len (struct_name)

            step = "1.1.2.2 Add global structs to buffer_prop"
            buffer_prop.append("//Global structs declaration\n")
            for struct_name in Proj.struct_list:
                struct_declaration = struct_name + "_struct"
                spaces_0 = " " * (max_spaces[0] - len(struct_declaration) + 1)
                spaces_1 = " " * (max_spaces[1] - len(struct_name) + 1)
                buffer_prop.append("{}{}\t{}{}\t".format(struct_declaration, spaces_0, struct_name, spaces_1))
                buffer_prop.append(" = {0};\n")
            buffer_prop.append("\n")

            step = "1.1.3 Write range_const list"
            buffer_prop.append("//range_const declaration\n")
            range_const_list = Proj.prop_list["sand_prop_range_t"]["range_const_list"]

            step = "1.1.3.1 Calc values len for align"
            for const_name in range_const_list:
                const = range_const_list[const_name]
                string = "static const " + const["type"]
                if len(string) > max_spaces[0]:
                    max_spaces[0] = len(string)
                if len(str(const["decl_name"])) > max_spaces[1]:
                    max_spaces[1] = len(str(const["decl_name"]))

            step = "1.1.3.2 Write range_const elements"
            for const_name in range_const_list:
                const = range_const_list[const_name]
                string = "static const " + const["type"]
                spaces_0 = " " * (max_spaces[0] - len(string) + 1)
                spaces_1 = " " * (max_spaces[1] - len(str(const["decl_name"])) + 1)
                buffer_prop.append("{}{}\t{}{} = {};\n".format(string, spaces_0, const["decl_name"], spaces_1, const["value"]))
            buffer_prop.append("\n")

            step = "1.1.4 Write prop lists "
            buffer_prop.append("//Property lists\n")
            for prop_name in Proj.prop_list:
                prop_list = Proj.prop_list[prop_name]

                step = "1.1.4.1 Calc headers len for align text"
                prop_param_list = list(sand_reg.sand_prop_list[prop_name].keys())
                prop_param_list.remove("header")
                header_param_list = list(sand_reg.sand_header_t.keys())
                if prop_name != "sand_prop_base_t":
                    max_spaces = [0] * (len(prop_param_list) + len(header_param_list) + 1)
                else:
                    max_spaces = [0] * (len(prop_param_list) + len(header_param_list))
                for header_param in header_param_list:
                    ind = header_param_list.index(header_param)
                    if len(header_param) > max_spaces[ind]:
                        max_spaces[ind] = len(header_param)
                for prop_param in prop_param_list:
                    ind = prop_param_list.index(prop_param) + len(header_param_list)
                    if len(prop_param) > max_spaces[ind]:
                        max_spaces[ind] = len(prop_param)
                if prop_name != "sand_prop_base_t":
                    ind = len(prop_param_list) + len(header_param_list)
                    for reg_name in prop_list["reg_list"]:
                        if len(reg_name) > max_spaces[ind]:
                            max_spaces[ind] = len(reg_name)


                step = "1.1.4.2 Calc regs values len for align text"
                for reg_name in prop_list["reg_list"]:
                    reg = prop_list["reg_list"][reg_name]
                    for header_param in header_param_list:
                        ind = header_param_list.index(header_param)
                        header = reg.prop_list[prop_name]["header"]
                        if len(header["header_t"][header_param]) > max_spaces[ind]:
                            max_spaces[ind] = len(header["header_t"][header_param])
                    for prop_param in prop_param_list:
                        ind = prop_param_list.index(prop_param) + len(header_param_list)
                        prop = reg.prop_list[prop_name][prop_param]
                        if "value" in prop:
                            if len(str(prop["value"])) > max_spaces[ind]:
                                max_spaces[ind] = len(str(prop["value"]))

                step = "1.1.5 Write prop_list to buffer_prop"
                prop_list_name = prop_name.replace("_t", "_list")
                prop_list_reg_num = prop_name.replace("_t", "").upper() + "_REG_NUM"
                buffer_prop.append("const {} {}[{}]".format(prop_name, prop_list_name, prop_list_reg_num))
                buffer_prop.append("={\n")

                step = "1.1.5.1 Write prop_list property parameters in comments"
                buffer_prop.append("//")
                for header_param in header_param_list:
                    ind = header_param_list.index(header_param)
                    spaces = " " * (max_spaces[ind] - len(header_param) + 1)
                    buffer_prop.append("{}{} \t".format(header_param, spaces))
                for prop_param in prop_param_list:
                    ind = prop_param_list.index(prop_param) + len(header_param_list)
                    spaces = " " * (max_spaces[ind] - len(prop_param) + 1)
                    buffer_prop.append("{}{} \t".format(prop_param, spaces))
                # Add reg name like comment to end of line
                if prop_name != "sand_prop_base_t":
                    ind = len(prop_param_list) + len(header_param_list)
                    spaces = " " * (max_spaces[ind] - len("reg_name") + 1)
                    buffer_prop.append("   reg_name")
                buffer_prop.append("\n")

                step = "1.1.5.2 Write prop_list values"
                for reg_name in prop_list["reg_list"]:
                    reg = prop_list["reg_list"][reg_name]
                    buffer_prop.append("{{")

                    step = "1.1.5.3 Write prop_list reg header values"
                    for header_param in header_param_list:
                        ind = header_param_list.index(header_param)
                        header_val = reg.prop_list[prop_name]["header"]["header_t"][header_param]
                        if header_param != header_param_list[-1]:
                            spaces = " " * (max_spaces[ind] - len(header_val))
                            buffer_prop.append("{},{} \t".format(header_val, spaces))
                        else:
                            spaces = " " * (max_spaces[ind] - len(header_val))
                            buffer_prop.append(header_val + "}," + spaces + " \t")

                    step = "1.1.5.4 Write prop_list reg prop values"
                    for prop_param in prop_param_list:
                        ind = prop_param_list.index(prop_param) + len(header_param_list)
                        if "value" in reg.prop_list[prop_name][prop_param]:
                            prop_val = reg.prop_list[prop_name][prop_param]["value"]
                        else:
                            prop_val = "error"
                        if prop_param != prop_param_list[-1]:
                            spaces = " " * (max_spaces[ind] - len(str(prop_val)))
                            buffer_prop.append("{},{} \t".format(prop_val, spaces))
                        else:
                            # Last prop_param (end of line)
                            if prop_name != "sand_prop_base_t":
                                ind = prop_param_list.index(prop_param) + len(header_param_list)
                                spaces = " " * (max_spaces[ind] - len(str(prop_val)) + 2)
                                buffer_prop.append("{}".format(prop_val) + "}," + spaces)
                                buffer_prop.append("// {}\n".format(reg_name))#"// {}\n".format(reg_name))
                            else:
                                buffer_prop.append("{}".format(prop_val) + "},\n")

                buffer_prop.append("};\n\n")
        except:
            Proj.errors["err_msg"].append("Error during adding property lists to " + file_name + ", step: " + step)
            Proj.errors["err_cnt"] += 1

    # 2. Rewrite file with new insertion from generator
    if prop_insert_find == True:

        file = open(Proj.path[file_name], 'w', encoding='UTF-8')
        line_index = 0
        # 2.1. Write file before insertion
        while line_index <= prop_insert_start:
            file.writelines(text_lines[line_index])
            line_index += 1
        # 2.2. Write property lists insertion
        for line in buffer_prop:
            file.writelines(line)
        # 2.3. Write file after insertion
        line_index = prop_insert_end
        while line_index < len(text_lines):
            file.writelines(text_lines[line_index])
            line_index += 1
        # 2.4. Close modified file
        file.close()

    return True

